# -*- coding: utf-8 -*-
"""
Created on Sun Jun 14 18:24:30 2020

@author: User
"""

import datetime
from openpyxl import load_workbook
import sys

if(len(sys.argv) != 2):
    print("Input Corect Format")
    print("TT_calc.py file_name.xlsx")
    sys.exit()

xlsx_name = sys.argv[1]

wb=load_workbook(xlsx_name)
sheet=wb.active


for row in sheet.rows:
    for cell in row:
        if type(cell.value) == str:
            if 'StartTime' in cell.value:
                row_start = cell.row
                column_start = cell.column

#print("row_start = ", row_start)
#print("column_start = ", column_start)



All = 0 
count = 0
for i in range(1, sheet.max_row):
    for j in range(column_start, column_start+1):
        start = (sheet.cell(row=i, column=j)).value
        end = (sheet.cell(row=i, column=(j+1))).value
        if start:
            if 'PASS' in (sheet.cell(row=i, column=(j-1)).value):
                All = All + (end - start).total_seconds()
                count = count + 1
        

print("avg TT = ", All/count)
#print("sheet_max_row = ", sheet.max_row)
#print("sheet_max_row = ", sheet.max_column)
